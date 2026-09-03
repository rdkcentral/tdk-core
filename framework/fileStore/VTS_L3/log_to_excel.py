##########################################################################
# If not stated otherwise in this file or this component's Licenses.txt
# file the following copyright and licenses apply:
#
# Copyright 2025 RDK Management
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#########################################################################

"""
Parses TDK/VTS test-run log files (dsAudio, dsVideo, dsHdmiIn, etc.) and
writes/updates an Excel workbook with:

  Sheet "Summary"
      | S.No | Module | No.Of.Tests | SUCCESS | FAILURE | RDK Issue |

  One sheet per module (e.g. "dsAudio")
      | S.No | Test Name | Status | Log Data | Jira | Remarks |

Calling the function again with a new log file:
  - appends new tests to the module sheet they belong to (creating the
    sheet/module row in Summary if it does not exist yet)
  - if a test with the same name already exists in that module's sheet,
    its row is OVERWRITTEN with the latest data
  - the Summary sheet counts are always recomputed from the actual rows
    present in each module sheet, so they stay accurate after every update
"""

import os
import re
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SUMMARY_SHEET = "Summary"
SUMMARY_HEADERS = ["S.No", "Module", "No.Of.Tests", "SUCCESS", "FAILURE", "RDK Issue"]
MODULE_HEADERS = ["S.No", "Test Name", "Status", "Log Data", "Jira", "Remarks"]

# ---- regex patterns -------------------------------------------------------

RUN_START_RE = re.compile(r'Running\s+(\S+)\.py\s+at')
RUN_END_RE = re.compile(r'Completed\s+(\S+)\.py\s+with exit code')
TEST_RESULT_RE = re.compile(r'TEST_RESULT\s*:\s*\[(PASSED|FAILED)\]')
STEP_LINE_RE = re.compile(r'(STEP_START\s*:.*|STEP_RESULT\s*:.*)$')
MODULE_FROM_NAME_RE = re.compile(r'^(.*?)_test\d+', re.IGNORECASE)

# ANSI escape sequences (color codes etc.) and other characters Excel's XML
# format refuses to store (control chars other than tab/newline/CR).
ANSI_ESCAPE_RE = re.compile(r'\x1b\[[0-9;]*[a-zA-Z]')
ILLEGAL_XML_CHARS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]'
)

# Excel's hard per-cell limit is 32,767 characters. Anything longer corrupts
# the file (Excel shows a "repaired records" dialog and blanks the cell).
EXCEL_CELL_CHAR_LIMIT = 32767
_TRUNCATION_NOTE = "\n\n... [TRUNCATED - log exceeded Excel's 32,767 char/cell limit] ...\n\n"


def _sanitize_for_excel(text):
    """Strip ANSI color codes and control characters that openpyxl/Excel
    cannot store in a cell, while keeping the text otherwise intact."""
    if not text:
        return text
    text = ANSI_ESCAPE_RE.sub("", text)
    text = ILLEGAL_XML_CHARS_RE.sub("", text)
    return text


def _clip_to_excel_limit(text):
    """If text is longer than Excel allows in one cell, keep the head and
    tail (most useful parts of a test log: setup + final result) and drop
    the middle, so the file never fails to open."""
    if not text or len(text) <= EXCEL_CELL_CHAR_LIMIT:
        return text
    budget = EXCEL_CELL_CHAR_LIMIT - len(_TRUNCATION_NOTE)
    head_len = budget * 2 // 3
    tail_len = budget - head_len
    return text[:head_len] + _TRUNCATION_NOTE + text[-tail_len:]


# Excel/openpyxl treats any string starting with one of these as a formula
# (or, for some, a number-like token) and tries to evaluate it instead of
# storing it as text. A log line starting with "=====" is a real-world case
# of this. Prefixing with a single quote is the standard Excel escape and
# keeps the text 100% intact otherwise.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _prevent_formula_interpretation(text):
    if text and text[0] in _FORMULA_TRIGGER_CHARS:
        return "'" + text
    return text


def _extract_module(test_case_name):
    """dsAudio_test01_Foo -> dsAudio (falls back to full name if pattern differs)."""
    m = MODULE_FROM_NAME_RE.match(test_case_name)
    return m.group(1) if m else test_case_name


def parse_log(log_path):
    """
    Reads a TDK log file and returns a list of dicts, one per test case found:
        {
            "module": "dsAudio",
            "test_name": "dsAudio_test01_EnableDisableAndVerifyAudioPortStatus",
            "status": "SUCCESS" | "FAILURE",
            "log_data": "<full block from 'Running ... at' to 'Completed ... exit code'>",
            "remarks": "<STEP_START / STEP_RESULT lines, in order>",
        }
    """
    with open(log_path, "r", errors="ignore") as f:
        lines = f.readlines()

    results = []
    i = 0
    n = len(lines)

    while i < n:
        start_match = RUN_START_RE.search(lines[i])
        if not start_match:
            i += 1
            continue

        test_case_name = start_match.group(1)
        block_lines = [lines[i]]
        j = i + 1
        end_found = False

        while j < n:
            block_lines.append(lines[j])
            end_match = RUN_END_RE.search(lines[j])
            if end_match and end_match.group(1) == test_case_name:
                end_found = True
                j += 1
                break
            j += 1

        block_text = "".join(block_lines).strip("\n")
        # start the captured log data at the word "Running" itself (per
        # spec), which also sidesteps the leading "=====" banner text
        running_idx = start_match.start()
        block_text = block_text[running_idx:]
        log_data = _clip_to_excel_limit(
            _prevent_formula_interpretation(_sanitize_for_excel(block_text))
        )

        # status: first PASSED/FAILED TEST_RESULT line inside the block
        status = ""
        result_match = TEST_RESULT_RE.search(log_data)
        if result_match:
            status = "SUCCESS" if result_match.group(1) == "PASSED" else "FAILURE"

        # remarks: every STEP_START / STEP_RESULT line, in original order
        remark_lines = []
        for line in block_lines:
            step_match = STEP_LINE_RE.search(line)
            if step_match:
                remark_lines.append(step_match.group(1).strip())
        remarks = _clip_to_excel_limit(
            _prevent_formula_interpretation(_sanitize_for_excel("\n".join(remark_lines)))
        )

        results.append({
            "module": _prevent_formula_interpretation(_extract_module(test_case_name)),
            "test_name": _prevent_formula_interpretation(test_case_name),
            "status": status,
            "log_data": log_data,
            "remarks": remarks,
        })

        i = j if end_found else i + 1

    return results


# ---- workbook helpers ------------------------------------------------------

def _safe_sheet_name(name):
    # Excel sheet names: max 31 chars, no []:*?/\
    cleaned = re.sub(r'[\[\]\:\*\?/\\]', "_", name)
    return cleaned[:31]


def _style_header(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _ensure_summary_sheet(wb):
    if SUMMARY_SHEET in wb.sheetnames:
        return wb[SUMMARY_SHEET]
    ws = wb.create_sheet(SUMMARY_SHEET, 0)
    _style_header(ws, SUMMARY_HEADERS)
    return ws


def _ensure_module_sheet(wb, module):
    sheet_name = _safe_sheet_name(module)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name], sheet_name
    ws = wb.create_sheet(sheet_name)
    _style_header(ws, MODULE_HEADERS)
    return ws, sheet_name


def _load_module_rows(ws):
    """Read existing module-sheet rows into an OrderedDict keyed by test name,
    preserving row order. Values: dict with status/log_data/jira/remarks."""
    rows = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        _, test_name, status, log_data, jira, remarks = (list(row) + [None] * 6)[:6]
        rows[test_name] = {
            "status": status or "",
            "log_data": log_data or "",
            "jira": jira or "",
            "remarks": remarks or "",
        }
    return rows


def _write_module_rows(ws, rows_dict):
    # wipe existing data rows (keep header row 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for idx, (test_name, data) in enumerate(rows_dict.items(), start=1):
        row_num = idx + 1
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=test_name)
        ws.cell(row=row_num, column=3, value=data.get("status", ""))
        ws.cell(row=row_num, column=4, value=data.get("log_data", ""))
        ws.cell(row=row_num, column=5, value=data.get("jira", "") or " ")
        ws.cell(row=row_num, column=6, value=data.get("remarks", ""))
        # Log Data / Remarks contain real embedded newlines, which Excel
        # auto-expands the row height for even with wrap_text off. Pin an
        # explicit short height so the default view stays compact; editing
        # a cell still shows its full multi-line content as normal.
        ws.row_dimensions[row_num].height = 15

    # reasonable column widths
    widths = {1: 6, 2: 45, 3: 10, 4: 60, 5: 12, 6: 60}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # a thin blocker column right after Remarks stops any viewer that
    # ignores wrap_text from spilling Remarks text into open space to the right
    blocker_col = len(MODULE_HEADERS) + 1
    ws.column_dimensions[get_column_letter(blocker_col)].width = 3
    for idx in range(1, len(rows_dict) + 1):
        ws.cell(row=idx + 1, column=blocker_col, value=" ")


def _recompute_summary(wb):
    summary_ws = wb[SUMMARY_SHEET]
    if summary_ws.max_row > 1:
        summary_ws.delete_rows(2, summary_ws.max_row - 1)

    module_sheets = [s for s in wb.sheetnames if s != SUMMARY_SHEET]

    for idx, sheet_name in enumerate(module_sheets, start=1):
        ws = wb[sheet_name]
        total = success = failure = rdk_issue = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None:
                continue
            status = (row[2] or "").strip().upper()
            total += 1
            if status == "SUCCESS":
                success += 1
            elif status == "FAILURE":
                failure += 1
            elif "RDK" in status:
                rdk_issue += 1

        summary_ws.cell(row=idx + 1, column=1, value=idx)
        summary_ws.cell(row=idx + 1, column=2, value=sheet_name)
        summary_ws.cell(row=idx + 1, column=3, value=total)
        summary_ws.cell(row=idx + 1, column=4, value=success)
        summary_ws.cell(row=idx + 1, column=5, value=failure)
        summary_ws.cell(row=idx + 1, column=6, value=rdk_issue)

    widths = {1: 6, 2: 22, 3: 14, 4: 10, 5: 10, 6: 12}
    for col_idx, width in widths.items():
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---- main entry point -------------------------------------------------------

def process_log_to_excel(log_path, excel_path="test_results.xlsx"):
    """
    Parses `log_path` and writes/updates `excel_path` with the Summary sheet
    and one sheet per module, following the rules described at the top of
    this file. Safe to call repeatedly with logs from different modules, or
    with a re-run of the same module (matching test names get overwritten).

    Returns the path to the saved workbook.
    """
    test_results = parse_log(log_path)
    if not test_results:
        raise ValueError(f"No test cases could be parsed from {log_path}")

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet

    _ensure_summary_sheet(wb)

    # group parsed tests by module, preserving first-seen order
    by_module = OrderedDict()
    for t in test_results:
        by_module.setdefault(t["module"], []).append(t)

    for module, tests in by_module.items():
        ws, _sheet_name = _ensure_module_sheet(wb, module)
        existing_rows = _load_module_rows(ws)
        for t in tests:
            existing_rows[t["test_name"]] = {
                "status": t["status"],
                "log_data": t["log_data"],
                "jira": existing_rows.get(t["test_name"], {}).get("jira", ""),
                "remarks": t["remarks"],
            }
        _write_module_rows(ws, existing_rows)

    _recompute_summary(wb)

    wb.save(excel_path)
    return excel_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python log_to_excel.py <log_file> [excel_file]")
        sys.exit(1)

    log_file = sys.argv[1]
    xlsx_file = sys.argv[2] if len(sys.argv) > 2 else "test_results.xlsx"
    saved_path = process_log_to_excel(log_file, xlsx_file)
    print(f"Saved: {saved_path}")
