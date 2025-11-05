import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font

def parse_qpa(qpa_path):
    results = []
    with open(qpa_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    xml_chunks = []
    inside = False
    for line in lines:
        if line.strip().startswith("<TestCaseResult"):
            inside = True
            xml_chunks = [line]
        elif line.strip().startswith("</TestCaseResult>"):
            xml_chunks.append(line)
            inside = False
            try:
                xml_str = "".join(xml_chunks)
                root = ET.fromstring(xml_str)
                case = root.attrib.get("CasePath", "")
                result_elem = root.find("Result")
                status = result_elem.attrib.get("StatusCode", "Unknown") if result_elem is not None else "Unknown"
                reason = result_elem.text.strip() if (result_elem is not None and result_elem.text) else ""
                results.append((case, status, reason))
            except Exception as e:
                print(f"Error parsing testcase: {e}")
        elif inside:
            xml_chunks.append(line)
    return results


def write_excel(qpa_dir, excel_out):
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Classification", "Passed", "Failed", "NotSupported"])

    bold_font = Font(bold=True)

    for fname in os.listdir(qpa_dir):
        if fname.endswith(".qpa"):
            classification = fname.split(".")[0]  # e.g., dEQP-VK.info
            classification = fname.rstrip("*.qpa").lstrip("dEQP-VK.")

            print("Parsing ",fname)
            print("Writing test results for class: ", classification)
            print("\n")
            results = parse_qpa(os.path.join(qpa_dir, fname))

            ws = wb.create_sheet(title=classification)
            ws.append(["TestCase", "Result", "Log"])

            # make header bold
            for cell in ws[1]:
                cell.font = bold_font

            passed = failed = not_supported = 0
            for case, status, reason in results:
                ws.append([case, status, reason])
                if status.lower() == "pass":
                    passed += 1
                elif status.lower() == "fail":
                    failed += 1
                elif status.lower() == "notsupported":
                    not_supported += 1

            summary.append([classification, passed, failed, not_supported])
            for cell in summary[1]:
                cell.font = bold_font

            # autosize cols
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # autosize summary too
    for col in summary.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        summary.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(excel_out)
    print(f"Results written to {excel_out}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python3 vulkan_cts.py <qpa_dir> <output.xlsx>")
        sys.exit(1)

    write_excel(sys.argv[1], sys.argv[2])
