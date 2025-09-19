##########################################################################
# If not stated otherwise in this file or this component's Licenses.txt
# file the following copyright and licenses apply:
#
# Copyright 2025 RDK Management
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#########################################################################

import xml.etree.ElementTree as ET
import sys
import requests
import argparse
import os
import re
from tabulate import tabulate
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Thin border style for outline
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==============================================================================
# Function      : fetch_documented_entries_from_md
# Description   : Fetches method and event names from the given Markdown documentation URL
#                 Differentiates methods and events based on naming
# Input         : URL of the Markdown file
# Output        : Set of documented method names and event names
# ==============================================================================
def fetch_documented_entries_from_md(md_url):
    try:
        response = requests.get(md_url)
        response.raise_for_status()
        content = response.text

        method_pattern = re.compile(r"\|\s*\[([a-zA-Z0-9_]+)\]\(#[^)]+\)")
        methods = set()
        events = set()

        for match in method_pattern.findall(content):
            name = match.strip()
            if name.lower().startswith("on"):
                events.add(name.lower())
            else:
                methods.add(name.lower())

        return methods, events

    except Exception as e:
        print(f"Error fetching or parsing methods/events from Markdown: {e}")
        return set(), set()

# ==============================================================================
# Function      : fetch_and_parse_xml
# Description   : Parses an XML file and returns the root of the XML tree
# Input         : File path to the XML
# Output        : Parsed XML root object or None if error
# ==============================================================================
def fetch_and_parse_xml(file_path):
    try:
        tree = ET.parse(file_path)
        return tree.getroot()
    except Exception as e:
        print(f"Error reading or parsing XML file: {e}")
        return None

# ==============================================================================
# Function      : extract_plugin_methods_and_events
# Description   : Extracts methods and events from the XML for the specified plugin service
#                 Handles prefix stripping if given
# Input         : XML root, service name, method/event prefix
# Output        : Dictionary of methods and events with their full names
# ==============================================================================
def extract_plugin_methods_and_events(xml_root, service_name, prefix):
    plugin = xml_root.find(f".//plugin[@serviceName='{service_name}']")
    if plugin is None:
        print(f"\nPlugin with serviceName '{service_name}' not found.")
        return {}, {}

    methods = {}
    events = {}

    for method in plugin.findall(".//method"):
        name_attr = method.get("name", "").strip()
        api_name = method.get("api", "").strip()
        short_name = name_attr.lower()

        if prefix and short_name.startswith(prefix.lower()):
            short_name = short_name[len(prefix):]

        method_key = api_name.lower() if api_name else short_name

        if method_key not in methods:
            methods[method_key] = set()
        methods[method_key].add(name_attr)

    for event in plugin.findall(".//event"):
        event_name = event.get("eventName", "").strip().lower()
        name_attr = event.get("name", "").strip()
        if event_name:
            if event_name not in events:
                events[event_name] = set()
            events[event_name].add(name_attr)

    return methods, events

# ==============================================================================
# Function      : map_methods_to_testcases
# Description   : Maps methods to test cases by scanning the testCase XML
#                 It checks both <testStep> and nested <subtestStep>.
# Input         : Path to testCase XML file
# Output        : Dictionary mapping method names to set of test case names
# ==============================================================================
def map_methods_to_testcases(testcase_xml_path):
    try:
        tree = ET.parse(testcase_xml_path)
        root = tree.getroot()
        method_name_to_testcases = {}

        for test_case in root.findall(".//testCase"):
            test_case_name = test_case.get("testCaseName", "").strip()

            for test_step in test_case.findall(".//testStep"):
                method_attr = test_step.get("method", "").strip().lower()
                if method_attr:
                    method_name_to_testcases.setdefault(method_attr, set()).add(test_case_name)

                # Check <subtestStep method="...">
                for sub_step in test_step.findall(".//subtestStep"):
                    sub_method_attr = sub_step.get("method", "").strip().lower()
                    if sub_method_attr:
                        method_name_to_testcases.setdefault(sub_method_attr, set()).add(test_case_name)

        return method_name_to_testcases
    
    except Exception as e:
        print(f"Error parsing test case XML: {e}")
        return {}

# ==============================================================================
# Function      : map_events_to_testcases
# Description   : Maps events to test cases from testCase XML
#                 Supports <event> in both <testStep> and <subtestStep>
# Input         : Path to testCase XML file
# Output        : Dictionary mapping event names to test case names
# ==============================================================================
def map_events_to_testcases(testcase_xml_path):
    try:
        tree = ET.parse(testcase_xml_path)
        root = tree.getroot()
        event_name_to_testcases = {}

        for test_case in root.findall(".//testCase"):
            test_case_name = test_case.get("testCaseName", "").strip()

            # Check event in <testStep>
            for test_step in test_case.findall(".//testStep"):
                event_attr = test_step.get("event", "")
                if event_attr:
                    event_name_to_testcases.setdefault(event_attr.lower().strip(), set()).add(test_case_name)

                # Also check inside <subtestStep>
                for sub_step in test_step.findall(".//subtestStep"):
                    sub_event_attr = sub_step.get("event", "")
                    if sub_event_attr:
                        event_name_to_testcases.setdefault(sub_event_attr.lower().strip(), set()).add(test_case_name)

        return event_name_to_testcases
    
    except Exception as e:
        print(f"Error parsing test case XML for events: {e}")
        return {}

# ==============================================================================
# Function      : compare_entries
# Description   : Compares documented entries (methods/events) with XML and test cases
#                 Generates Excel sheet with status (Automated/Not Automated/Undocumented)
# Input         : Documented set, XML dictionary, testcase map, section title, output path
# Output        : Summary dictionary with counts
# ==============================================================================
def compare_entries(doc_entries, xml_entries, testcase_map, title="Method", output_excel_path="api_comparison.xlsx", global_unique_testcases=None):
    all_names = sorted(doc_entries | set(xml_entries.keys()))
    table_data = []

    automated = 0
    not_automated = 0
    undocumented = 0

    section_unique_testcases = set()

    for i, name in enumerate(all_names, start=1):
        in_doc = name in doc_entries
        in_xml = name in xml_entries

        api_names = xml_entries.get(name, set())
        testcase_list = set()
        for full_name in api_names:
            testcase_list.update(testcase_map.get(full_name.lower(), set()))

        if global_unique_testcases is not None:
            global_unique_testcases.update(testcase_list)

        testcase_names = "\n".join(sorted(testcase_list)) if testcase_list else ""
        testcase_count = len(testcase_list)

        if in_doc and in_xml:
            status = "Automated"
            automated += 1
        elif in_doc and not in_xml:
            status = "Not Automated"
            not_automated += 1
        elif not in_doc and in_xml:
            status = "Undocumented (Only in XML)"
            undocumented += 1
        else:
            status = "Unknown"

        table_data.append([i, name, status, testcase_names, testcase_count])

    if not table_data:
        print(f"\nNo {title}s found to compare")
        return None

    print(f"\n{title} Comparison:")
    print(tabulate(
        table_data,
        headers=["S.No", f"{title} Name", "Status", "Test Case Name", "Test Case Count"],
        tablefmt="fancy_grid",
        disable_numparse=True
    ))

    summary = {
        'total_docs': len(doc_entries),
        'automated': automated,
        'not_automated': not_automated,
        'undocumented': undocumented,
    }

    df = pd.DataFrame(table_data, columns=["S.No", f"{title} Name", "Status", "Test Case Name", "Test Case Count"])

    try:
        with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a' if os.path.exists(output_excel_path) else 'w') as writer:
            df.to_excel(writer, index=False, sheet_name=f"{title}s")

        wb = load_workbook(output_excel_path)
        ws = wb[f"{title}s"]

        max_testcase_width = 100
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            header = col_cells[0].value
            if header == "Test Case Name":
                ws.column_dimensions[col_letter].width = max_testcase_width
            else:
                max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
                ws.column_dimensions[col_letter].width = max_length + 2

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Center headers
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply outline border to the entire table range
        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border

        wb.save(output_excel_path)

    except Exception as e:
        print(f"Failed to write {title} to Excel: {e}")

    return summary

# ==============================================================================
# Function      : write_summary_sheet_simple_centered
# Description   : Writes a formatted Summary sheet to the Excel with hyperlinks and alignment
# Input         : Excel path, method summary, event summary, total unique test case count
# Output        : Writes to Excel
# ==============================================================================
def write_summary_sheet_simple_centered(output_excel_path, method_summary, event_summary, unique_testcase_count):
    if os.path.exists(output_excel_path):
        wb = load_workbook(output_excel_path)
    else:
        from openpyxl import Workbook
        wb = Workbook()

    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        wb.remove(ws)
    ws = wb.create_sheet('Summary', 0)

    start_col = 4  # Column D
    start_row = 10  # Row 10

    headers = ["S.No", "Section", "Total API in Docs", "Automated", "Not Automated", "Undocumented", "Total Unique Test Cases"]
    data = []
    row_index = 1

    if method_summary:
        data.append([
            row_index,
            "Methods",
            method_summary['total_docs'],
            method_summary['automated'],
            method_summary['not_automated'],
            method_summary['undocumented'],
            '—'
        ])
        row_index += 1

    if event_summary:
        data.append([
            row_index,
            "Events",
            event_summary['total_docs'],
            event_summary['automated'],
            event_summary['not_automated'],
            event_summary['undocumented'],
            '—'
        ])
        row_index += 1

    # Add total unique test cases row
    data.append(["", "", "", "", "", "", unique_testcase_count])

    col_widths = [8, 15, 18, 13, 18, 15, 22]
    for i, width in enumerate(col_widths, start=start_col):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Write headers
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + col_offset, value=header)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = Font(bold=True)

    # Write data rows
    for row_offset, row_data in enumerate(data, start=1):
        row_num = start_row + row_offset
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(row=row_num, column=start_col + col_offset, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Only add hyperlink for Methods or Events (column index 1)
            if col_offset == 1 and value in ['Methods', 'Events']:
                target_sheet = value
                cell.hyperlink = f"#{target_sheet}!A1"
                cell.font = Font(color='0000FF', underline='single')

    # Apply borders
    max_row = start_row + len(data)
    max_col = start_col + len(headers) - 1
    for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=start_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    wb.save(output_excel_path)

# ==============================================================================
# Function      : print_summary_terminal
# Description   : Prints a tabular summary of methods and events comparison in the terminal
# Input         : Method summary dict, event summary dict, unique test case count
# Output        : None (console output)
# ==============================================================================
def print_summary_terminal(method_summary, event_summary, unique_testcase_count):
    headers = ["S.No", "Section", "Total API in Docs", "Automated", "Not Automated", "Undocumented", "Total Unique Test Cases"]
    rows = []
    row_index = 1
    if method_summary:
        rows.append([row_index, "Methods", method_summary['total_docs'], method_summary['automated'], method_summary['not_automated'], method_summary['undocumented'], '—'])
        row_index += 1
    if event_summary:
        rows.append([row_index, "Events", event_summary['total_docs'], event_summary['automated'], event_summary['not_automated'], event_summary['undocumented'], '—'])
        row_index += 1

    rows.append(["", "", "", "", "", "", unique_testcase_count])

    print("\nSummary:")
    print(tabulate(rows, headers=headers, tablefmt="fancy_grid"))

# ==============================================================================
# Function      : find_unmatched_testcases
# Description   : Finds test cases that are not mapped to any valid method or event
# Input         : Path to testcase XML, all known valid method and event names
# Output        : Set of unmatched test case names
# ==============================================================================
def find_unmatched_testcases(testcase_xml_path, all_valid_methods, all_valid_events):
    try:
        tree = ET.parse(testcase_xml_path)
        root = tree.getroot()

        unmatched_testcases = set()

        for test_case in root.findall(".//testCase"):
            test_case_name = test_case.get("testCaseName", "").strip()
            found_valid = False

            # Check <testStep>
            for test_step in test_case.findall(".//testStep"):
                method = test_step.get("method", "").strip().lower()
                event = test_step.get("event", "").strip().lower()

                if method and method in all_valid_methods:
                    found_valid = True
                if event and event in all_valid_events:
                    found_valid = True

                # Check <subtestStep>
                for sub_step in test_step.findall(".//subtestStep"):
                    sub_method = sub_step.get("method", "").strip().lower()
                    sub_event = sub_step.get("event", "").strip().lower()

                    if sub_method and sub_method in all_valid_methods:
                        found_valid = True
                    if sub_event and sub_event in all_valid_events:
                        found_valid = True

            if not found_valid:
                unmatched_testcases.add(test_case_name)

        return unmatched_testcases

    except Exception as e:
        print(f"Error finding unmatched test cases: {e}")
        return set()

# ==============================================================================
# Function      : write_unmatched_testcases_to_excel
# Description   : Writes unmatched test cases to a separate sheet in the Excel output
# Input         : Excel path, set of unmatched test cases
# Output        : Writes to Excel
# ==============================================================================
def write_unmatched_testcases_to_excel(output_excel_path, unmatched_testcases):
    try:
        if os.path.exists(output_excel_path):
            wb = load_workbook(output_excel_path)
        else:
            from openpyxl import Workbook
            wb = Workbook()

        sheet_name = "Unmatched TestCases"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)

        ws = wb.create_sheet(sheet_name)

        # Header
        headers = ["S.No", "Test Case Name"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.border = thin_border

        if unmatched_testcases:
            for idx, tc_name in enumerate(sorted(unmatched_testcases), start=1):
                ws.cell(row=idx + 1, column=1, value=idx).border = thin_border
                ws.cell(row=idx + 1, column=2, value=tc_name).border = thin_border
        else:
            # Show success message if no unmatched cases
            ws.cell(row=2, column=1, value="1").border = thin_border
            ws.cell(row=2, column=2, value="All test cases are mapped to methods or events").border = thin_border

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60

        wb.save(output_excel_path)
    except Exception as e:
        print(f"Error writing unmatched test cases to Excel: {e}")

# ==============================================================================
# Function      : print_custom_usage
# Description   : Prints example usage and help message when arguments are missing
# Input         : None
# Output        : Printed to console
# ==============================================================================
def print_custom_usage():
    print("""
Usage
TestcaseGapFinder.py --md-url <MD_URL> --xml-file <XML_FILE> --service-name <SERVICE_NAME>
           [--prefix PREFIX] --testcase-xml TESTCASE_XML [--output-excel OUTPUT_EXCEL]
Note: The following arguments are required --md-url, --xml-file, --service-name, --testcase-xml

TestcaseGapFinder has been tested with various RDK Services documentation and is working as expected
Verified plugins include HDCPProfile, System, UserPreference, FrameRate, DeviceDiagnostics,
RDKShell, Monitor, and PersistentStore
""")

# ==============================================================================
# Function      : main
# Description   : Orchestrates the flow
#                 - Parses arguments
#                 - Loads and compares methods/events
#                 - Generates Excel reports and prints summary
# Input         : None (uses sys.argv)
# Output        : Terminal output + Excel report
# ==============================================================================
def main():
    if len(sys.argv) == 1:
        print_custom_usage()
        return

    parser = argparse.ArgumentParser(description="Compare documented and XML plugin methods/events.")
    parser.add_argument("--md-url", required=True)
    parser.add_argument("--xml-file", required=True)
    parser.add_argument("--service-name", required=True)
    parser.add_argument("--prefix", default="")
    parser.add_argument("--testcase-xml", required=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"api_comparison_{timestamp}.xlsx"
    parser.add_argument("--output-excel", default=default_filename)

    args = parser.parse_args()

    if not os.path.exists(args.xml_file):
        print(f"XML file not found: {args.xml_file}")
        return
    if not os.path.exists(args.testcase_xml):
        print(f"Test case XML file not found: {args.testcase_xml}")
        return

    documented_methods, documented_events = fetch_documented_entries_from_md(args.md_url)
    xml_root = fetch_and_parse_xml(args.xml_file)
    method_testcase_map = map_methods_to_testcases(args.testcase_xml)
    event_testcase_map = map_events_to_testcases(args.testcase_xml)

    if xml_root:
        xml_methods, xml_events = extract_plugin_methods_and_events(xml_root, args.service_name, args.prefix)

        global_unique_testcases = set()

        method_summary = compare_entries(
            documented_methods,
            xml_methods,
            method_testcase_map,
            title="Method",
            output_excel_path=args.output_excel,
            global_unique_testcases=global_unique_testcases
        )

        event_summary = compare_entries(
            documented_events,
            xml_events,
            event_testcase_map,
            title="Event",
            output_excel_path=args.output_excel,
            global_unique_testcases=global_unique_testcases
        )

        unique_count = len(global_unique_testcases)

        write_summary_sheet_simple_centered(args.output_excel, method_summary, event_summary, unique_count)
        print_summary_terminal(method_summary, event_summary, unique_count)

        # Get all test case names from test case XML
        all_testcases = set()
        try:
            tree = ET.parse(args.testcase_xml)
            root = tree.getroot()
            for test_case in root.findall(".//testCase"):
                name = test_case.get("testCaseName", "").strip()
                if name:
                    all_testcases.add(name)
        except Exception as e:
            print(f"Error reading test case XML for unmatched detection: {e}")

        unmatched_testcases = all_testcases - global_unique_testcases

        if unmatched_testcases:
            print("\nUnused Test Cases (Not matched to any method/event):")
            for name in sorted(unmatched_testcases):
                print(f"  - {name}")
        else:
            print("\nAll test cases are mapped to methods or events.")

        write_unmatched_testcases_to_excel(args.output_excel, unmatched_testcases)

        print(f"\nExcel file generated successfully: {args.output_excel}")

if __name__ == "__main__":
    main()
